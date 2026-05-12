#!/usr/bin/env python3
"""
import-us-benchmarks.py

Reads Nasdaq100 and SP500 XLSX files, computes monthly decimal returns,
and writes them to KV (benchmarks:green).

Operations:
  - Adds category field to ALL existing benchmarks
  - Updates bm-sp500 (placeholder → full history from XLSX)
  - Adds bm-nasdaq100 as new entry

Usage (from project root):
  python3 scripts/import-us-benchmarks.py [--env .env.production.local]
"""

import json
import os
import sys
import urllib.request
import urllib.error
from pathlib import Path
from collections import defaultdict
from datetime import datetime, timezone

# Force UTF-8 output on Windows
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# ── Config ────────────────────────────────────────────────────────────────────

DEFAULT_SOURCE_DIR = r"C:\Users\Agam\Desktop\מעקב קרנות\מדדים היסטורים"

XLSX_MAP = {
    "bm-nasdaq100": {
        "file": "Nasdaq100_monthly_2014_2026.xlsx",
        "sheet": "Nasdaq-100 Monthly",
        "name": "Nasdaq 100",
        "currency": "USD",
    },
    "bm-sp500": {
        "file": "SP500_monthly_2014_2026.xlsx",
        "sheet": "S&P-500 Monthly",
        "name": "S&P 500",
        "currency": "USD",
    },
}

CATEGORY_MAP = {
    "bm-ta125":          "מדדי מניות ישראל",
    "bm-sme60":          "מדדי מניות ישראל",
    "bm-telbond-maagar": 'מדדי אג"ח ישראל',
    "bm-agach-klali":    'מדדי אג"ח ישראל',
    "bm-nasdaq100":      'מדדי חו"ל',
    "bm-sp500":          'מדדי חו"ל',
}

KV_KEY = "benchmarks:green"
CURRENT_YEAR = 2026  # partial — no yearly return computed


# ── Env loader ────────────────────────────────────────────────────────────────

def load_env(env_file: str) -> None:
    p = Path(env_file)
    if not p.exists():
        # Try relative to main project root (one level up from scripts/)
        alt = Path(__file__).parent.parent / ".env.production.local"
        if alt.exists():
            p = alt
        else:
            print(f"ERROR: env file not found: {env_file}")
            print("Run: vercel env pull .env.production.local")
            sys.exit(1)

    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        eq = line.find("=")
        if eq < 0:
            continue
        key = line[:eq].strip()
        val = line[eq + 1:].strip().strip('"\'')
        os.environ[key] = val


# ── XLSX reader ───────────────────────────────────────────────────────────────

def read_xlsx_monthly_returns(bm_id: str, source_dir: str) -> dict:
    from openpyxl import load_workbook

    config = XLSX_MAP[bm_id]
    file_path = Path(source_dir) / config["file"]

    if not file_path.exists():
        raise FileNotFoundError(f"Not found: {file_path}")

    wb = load_workbook(str(file_path), data_only=True)

    if config["sheet"] in wb.sheetnames:
        ws = wb[config["sheet"]]
    else:
        ws = wb.active
        print(f"  WARNING: sheet '{config['sheet']}' not found, using '{ws.title}'")

    rows = list(ws.values)
    if not rows:
        raise ValueError(f"Empty workbook: {file_path}")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    date_col = headers.index("Date")
    close_col = headers.index("Close")

    # Collect (month_key, close) pairs in order
    closes = []
    for row in rows[1:]:
        date_val = row[date_col]
        close_val = row[close_col]
        if date_val is None or close_val is None:
            continue

        if isinstance(date_val, str):
            month_key = date_val[:7]
        else:
            month_key = date_val.strftime("%Y-%m")

        close = float(close_val)
        closes.append((month_key, close))

    # Compute monthly returns (decimal), skip first month (no prev close)
    monthly_returns = {}
    for i in range(1, len(closes)):
        month_key, close = closes[i]
        _, prev_close = closes[i - 1]
        r = round(close / prev_close - 1, 6)
        monthly_returns[month_key] = r

    return monthly_returns


# ── Yearly / returns object ───────────────────────────────────────────────────

def compute_yearly(mr: dict) -> dict:
    by_year = defaultdict(list)
    for month, ret in sorted(mr.items()):
        year = int(month[:4])
        by_year[year].append(ret)

    result = {}
    for year, rets in sorted(by_year.items()):
        if year >= CURRENT_YEAR:
            continue
        compound = 1.0
        for r in rets:
            compound *= 1.0 + r
        result[str(year)] = round(compound - 1, 4)
    return result


def build_returns_obj(mr: dict) -> dict:
    yearly = compute_yearly(mr)

    # YTD = compound of all CURRENT_YEAR months present
    ytd_rets = [v for m, v in sorted(mr.items()) if m.startswith(f"{CURRENT_YEAR}-")]
    ytd = None
    if ytd_rets:
        c = 1.0
        for r in ytd_rets:
            c *= 1.0 + r
        ytd = round(c - 1, 4)

    return {
        "ytd2026": ytd,
        "y2025": yearly.get("2025"),
        "y2024": yearly.get("2024"),
        "y2023": yearly.get("2023"),
        "y2022": yearly.get("2022"),
        "y2021": yearly.get("2021"),
        "y2020": yearly.get("2020"),
        "y2019": yearly.get("2019"),
    }


# ── KV helpers ────────────────────────────────────────────────────────────────

def _kv_request(body: list) -> object:
    url = os.environ["KV_REST_API_URL"]
    token = os.environ["KV_REST_API_TOKEN"]
    payload = json.dumps(body).encode("utf-8")
    req = urllib.request.Request(
        url + "/",
        data=payload,
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read().decode("utf-8"))


def kv_get(key: str) -> object:
    result = _kv_request(["GET", key])["result"]
    return json.loads(result) if result else None


def kv_set(key: str, value: object) -> None:
    result = _kv_request(["SET", key, json.dumps(value, ensure_ascii=False)])["result"]
    if result != "OK":
        raise RuntimeError(f"KV SET failed: {result}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    # Parse args
    env_file = ".env.production.local"
    source_dir = DEFAULT_SOURCE_DIR
    for i, arg in enumerate(sys.argv[1:]):
        if arg == "--env" and i + 1 < len(sys.argv[1:]):
            env_file = sys.argv[i + 2]
        elif arg == "--source" and i + 1 < len(sys.argv[1:]):
            source_dir = sys.argv[i + 2]

    load_env(env_file)

    if not os.environ.get("KV_REST_API_URL") or not os.environ.get("KV_REST_API_TOKEN"):
        print("ERROR: KV_REST_API_URL and KV_REST_API_TOKEN must be set")
        sys.exit(1)

    print("=== US Benchmark Import ===\n")
    now_iso = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")

    # ── Step 1: Process XLSX files ─────────────────────────────────────────
    xlsx_data = {}
    for bm_id in XLSX_MAP:
        print(f"Reading {bm_id} from XLSX...")
        mr = read_xlsx_monthly_returns(bm_id, source_dir)
        months = sorted(mr.keys())
        returns_obj = build_returns_obj(mr)
        print(f"  {len(months)} months: {months[0]} → {months[-1]}")
        print(f"  ytd2026={returns_obj['ytd2026']:.4f}  y2025={returns_obj['y2025']:.4f}  y2024={returns_obj['y2024']:.4f}")

        # Range validation (monthly returns should be in [-0.5, 0.5])
        outliers = [(m, v) for m, v in mr.items() if abs(v) > 0.5]
        if outliers:
            print(f"  WARNING: {len(outliers)} months outside ±50% range: {outliers[:3]}")

        xlsx_data[bm_id] = {"monthlyReturns": mr, "returns": returns_obj}

    # ── Step 2: Read current KV ────────────────────────────────────────────
    print("\nReading current KV...")
    benchmarks = kv_get(KV_KEY)
    if not isinstance(benchmarks, list):
        raise RuntimeError("KV benchmarks:green is not an array")

    print(f"KV has {len(benchmarks)} benchmarks: {[b['id'] for b in benchmarks]}")

    # ── Step 3: Add category to all existing ──────────────────────────────
    print("\nAdding category field to existing benchmarks:")
    for bm in benchmarks:
        cat = CATEGORY_MAP.get(bm["id"], "לא מסווג")
        old_cat = bm.get("category", "MISSING")
        bm["category"] = cat
        if old_cat != cat:
            print(f"  {bm['id']}: '{old_cat}' → '{cat}'")
        else:
            print(f"  {bm['id']}: already '{cat}'")

    # ── Step 4: Update/add US benchmarks ──────────────────────────────────
    existing_ids = {b["id"] for b in benchmarks}

    for bm_id, data in xlsx_data.items():
        config = XLSX_MAP[bm_id]
        mr = data["monthlyReturns"]
        months = sorted(mr.keys())

        if bm_id in existing_ids:
            bm = next(b for b in benchmarks if b["id"] == bm_id)
            old_count = len(bm.get("monthlyReturns") or {})
            bm["monthlyReturns"] = mr
            bm["returns"] = data["returns"]
            bm["category"] = CATEGORY_MAP[bm_id]
            bm["source"] = "Investing.com"
            bm["lastUpdated"] = now_iso
            print(f"\nUPDATE {bm_id} ({bm['name']}): {old_count} months → {len(months)} months ({months[0]} → {months[-1]})")
        else:
            new_bm = {
                "id": bm_id,
                "name": config["name"],
                "currency": config["currency"],
                "category": CATEGORY_MAP[bm_id],
                "returns": data["returns"],
                "monthlyReturns": mr,
                "active": True,
                "source": "Investing.com",
                "lastUpdated": now_iso,
            }
            benchmarks.append(new_bm)
            print(f"\nADD {bm_id} ({config['name']}): {len(months)} months ({months[0]} → {months[-1]})")

    # ── Step 5: Write to KV ────────────────────────────────────────────────
    print(f"\nWriting {len(benchmarks)} benchmarks to KV...")
    kv_set(KV_KEY, benchmarks)
    print("Write OK")

    # ── Step 6: Validate (read-back) ───────────────────────────────────────
    print("\n=== Validation (read-back) ===")
    final = kv_get(KV_KEY)
    if not isinstance(final, list):
        raise RuntimeError("Validation failed: KV returned non-array")

    print(f"{'ID':<25} {'Name':<22} {'Months':>6}  {'First':>7} → {'Last':<7}  Category")
    print("-" * 100)
    for b in final:
        mr = b.get("monthlyReturns") or {}
        months = sorted(mr.keys())
        first = months[0] if months else "none"
        last = months[-1] if months else "none"
        cat = b.get("category", "MISSING")
        print(f"  {b['id']:<23} {b.get('name',''):<22} {len(months):>6}  {first:>7} → {last:<7}  {cat}")

    print("\nDone.")


if __name__ == "__main__":
    main()
