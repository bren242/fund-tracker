"""
load-benchmark-history.py
=========================
טוען נתונים היסטוריים ל-benchmarks:green ב-Vercel KV.

קבצי קלט (C:/Users/Agam/Desktop/מעקב קרנות/מדדים היסטורים/):
  TA125_monthly_2014_2026.xlsx    — Date, Close (חישוב תשואה month-over-month)
  מאגר_monthly.xlsx               — Date, Close, Monthly_Return
  SME60_monthly.xlsx              — Date, Close, Monthly_Return
  קונצרני_כללי_monthly.xlsx       — Date, Close, Monthly_Return

פעולות:
  1. קרא קבצים → monthlyReturns (סינון >= START_FROM)
  2. חשב stdDev (sample) לכל מדד
  3. חשב ytd2026 מחדש (compound) אם יש נתוני 2026
  4. קרא benchmarks:green מ-KV (fallback: קובץ local)
  5. Preview — מה עומד להשתנות
  6. [אישור] → כתוב ל-KV
  7. אימות

הרץ: python scripts/load-benchmark-history.py
"""

import openpyxl
import json
import urllib.request
import urllib.error
import math
import os
import sys
import pprint
from datetime import datetime

# ── Config ────────────────────────────────────────────────────────────────────

KV_URL   = "https://upright-asp-87838.upstash.io"
KV_TOKEN = "gQAAAAAAAVceAAIncDEwNjZlZTFiMTQ3Y2I0YjRhODMwYTkzZWNhZjFjZDIxZnAxODc4Mzg"

DATA_DIR    = "C:/Users/Agam/Desktop/מעקב קרנות/מדדים היסטורים"
LOCAL_BM    = "C:/Users/Agam/Desktop/מעקב קרנות/fund-tracker/data/green/benchmarks.json"
START_FROM  = "2020-01"  # סינון — רק ממועד זה ואילך

# מיפוי קובץ → id קיים ב-KV
FILE_TO_BM_ID = {
    "TA125_monthly_2014_2026.xlsx":  "bm-ta125",
    "מאגר_monthly.xlsx":             "bm-telbond-maagar",
    "קונצרני_כללי_monthly.xlsx":     "bm-agach-klali",
    # SME60 — חדש, ייווצר אם לא קיים
    "SME60_monthly.xlsx":            "bm-sme60",
}

SME60_NEW_BENCHMARK = {
    "id": "bm-sme60",
    "name": "SME 60",
    "currency": "ILS",
    "returns": {
        "ytd2026": None,
        "y2025": None, "y2024": None, "y2023": None,
        "y2022": None, "y2021": None, "y2020": None, "y2019": None,
    },
    "monthlyReturns": {},
    "active": True,
}

# ── KV helpers ────────────────────────────────────────────────────────────────

def kv_get(key):
    """קרא ערך מ-KV. מחזיר None אם לא קיים."""
    url = f"{KV_URL}/get/{urllib.parse.quote(key, safe='')}"
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {KV_TOKEN}"})
    try:
        with urllib.request.urlopen(req) as resp:
            data = json.loads(resp.read())
            result = data.get("result")
            if result is None:
                return None
            return json.loads(result) if isinstance(result, str) else result
    except urllib.error.HTTPError as e:
        print(f"  KV GET error {e.code}: {e.reason}")
        return None

def kv_set(key, value):
    """כתוב ערך ל-KV."""
    import urllib.parse
    url = f"{KV_URL}/set/{urllib.parse.quote(key, safe='')}"
    body = json.dumps(value).encode()
    req = urllib.request.Request(
        url, data=body, method="POST",
        headers={
            "Authorization": f"Bearer {KV_TOKEN}",
            "Content-Type": "application/json",
        }
    )
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read())

# ── Date utils ────────────────────────────────────────────────────────────────

def to_yyyymm(date_val) -> str | None:
    """ממיר ערך תאריך (string/datetime/date) לפורמט YYYY-MM."""
    if date_val is None:
        return None
    if isinstance(date_val, str):
        s = date_val.strip()
        if len(s) >= 7 and s[4] == "-":
            return s[:7]
        return None
    # datetime / date object
    return date_val.strftime("%Y-%m")

# ── Excel readers ─────────────────────────────────────────────────────────────

def read_ta125(filename: str) -> dict:
    """
    קובץ עם Close בלבד.
    מחשב תשואה חודשית: (Close[i] / Close[i-1]) - 1
    מחזיר: { "YYYY-MM": return_value, ... }
    """
    path = os.path.join(DATA_DIR, filename)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    rows = []
    for r in range(2, ws.max_row + 1):
        date_val = ws.cell(r, 1).value
        close    = ws.cell(r, 2).value
        key      = to_yyyymm(date_val)
        if key and close is not None:
            rows.append((key, float(close)))

    # חישוב תשואה month-over-month
    result = {}
    for i in range(1, len(rows)):
        ym_prev, close_prev = rows[i - 1]
        ym_curr, close_curr = rows[i]
        if close_prev and close_prev != 0:
            ret = (close_curr / close_prev) - 1
            result[ym_curr] = round(ret, 6)

    return result

def read_with_returns(filename: str) -> dict:
    """
    קובץ עם עמודת Monthly_Return מוכנה.
    מחזיר: { "YYYY-MM": return_value, ... }
    """
    path = os.path.join(DATA_DIR, filename)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    result = {}
    for r in range(2, ws.max_row + 1):
        date_val = ws.cell(r, 1).value
        ret_val  = ws.cell(r, 3).value  # עמודה C = Monthly_Return
        key      = to_yyyymm(date_val)
        if key and ret_val is not None:
            result[key] = round(float(ret_val), 6)

    return result

# ── Statistics ────────────────────────────────────────────────────────────────

def compute_stddev(returns: dict) -> float | None:
    """
    חישוב סטיית תקן חודשית — sample (÷ N-1).
    מחזיר None אם פחות מ-12 תצפיות.
    """
    vals = list(returns.values())
    n = len(vals)
    if n < 12:
        return None
    mean = sum(vals) / n
    variance = sum((v - mean) ** 2 for v in vals) / (n - 1)
    return round(math.sqrt(variance), 6)

def compute_ytd(monthly: dict, year: int) -> float | None:
    """
    מחשב compound YTD לשנה נתונה מתשואות חודשיות.
    """
    prefix = f"{year}-"
    months = sorted([(k, v) for k, v in monthly.items() if k.startswith(prefix)])
    if not months:
        return None
    cumulative = 1.0
    for _, v in months:
        cumulative *= (1 + v)
    return round(cumulative - 1, 6)

def compute_annual(monthly: dict, year: int) -> float | None:
    """
    מחשב תשואה שנתית compound אם יש 12 חודשים.
    """
    prefix = f"{year}-"
    months = sorted([(k, v) for k, v in monthly.items() if k.startswith(prefix)])
    if len(months) < 12:
        return None
    cumulative = 1.0
    for _, v in months:
        cumulative *= (1 + v)
    return round(cumulative - 1, 6)

# ── Filter ────────────────────────────────────────────────────────────────────

def filter_from(monthly: dict, start: str) -> dict:
    """סינון — רק רשומות >= start (YYYY-MM)."""
    return {k: v for k, v in monthly.items() if k >= start}

# ── Main ──────────────────────────────────────────────────────────────────────

import urllib.parse

def main():
    print("=" * 60)
    print("load-benchmark-history.py")
    print("=" * 60)

    # ── שלב 1: קרא קבצים ──────────────────────────────────────
    print("\n📂 קורא קבצי Excel...")
    raw_data = {}

    for filename, bm_id in FILE_TO_BM_ID.items():
        path = os.path.join(DATA_DIR, filename)
        if not os.path.exists(path):
            print(f"  ⚠️  קובץ לא נמצא: {filename}")
            continue
        if filename == "TA125_monthly_2014_2026.xlsx":
            monthly = read_ta125(filename)
        else:
            monthly = read_with_returns(filename)
        monthly_filtered = filter_from(monthly, START_FROM)
        raw_data[bm_id] = {
            "all_monthly":      monthly,
            "monthly_filtered": monthly_filtered,
            "stddev":           compute_stddev(monthly_filtered),
            "ytd2026":          compute_ytd(monthly_filtered, 2026),
        }
        keys = sorted(monthly_filtered.keys())
        print(f"  ✓ {filename}")
        print(f"     {bm_id} | {len(monthly_filtered)} חודשים | {keys[0] if keys else '?'} – {keys[-1] if keys else '?'}")
        print(f"     stdDev={raw_data[bm_id]['stddev']} | ytd2026={raw_data[bm_id]['ytd2026']}")

    # ── שלב 2: קרא benchmarks:green מ-KV ──────────────────────
    print("\n🔑 קורא benchmarks:green מ-KV...")
    existing = kv_get("benchmarks:green")
    if existing is None:
        print("  לא נמצא ב-KV — טוען מקובץ local")
        with open(LOCAL_BM, encoding="utf-8") as f:
            existing = json.load(f)
        print(f"  ✓ נטען מ-{LOCAL_BM} ({len(existing)} מדדים)")
    else:
        print(f"  ✓ נמצא ב-KV ({len(existing)} מדדים)")

    existing_ids = {bm["id"] for bm in existing}
    print(f"  מדדים קיימים: {', '.join(sorted(existing_ids))}")

    # ── שלב 3: בנה מבנה מעודכן ──────────────────────────────────
    updated = []
    for bm in existing:
        bm_id = bm["id"]
        if bm_id in raw_data:
            d = raw_data[bm_id]
            # merge: שמור כל שדה קיים, עדכן רק monthlyReturns + stdDev + ytd2026
            new_monthly = {
                **(bm.get("monthlyReturns") or {}),
                **d["monthly_filtered"],          # החדש דורס ישן לפי key
            }
            new_bm = {
                **bm,
                "monthlyReturns": new_monthly,
                "stdDev": d["stddev"],
            }
            # עדכן ytd2026 אם יש נתוני 2026
            if d["ytd2026"] is not None:
                new_bm["returns"] = {**bm.get("returns", {}), "ytd2026": d["ytd2026"]}
            updated.append(new_bm)
        else:
            updated.append(bm)   # לא נגעים

    # הוסף SME60 אם חדש
    if "bm-sme60" not in existing_ids and "bm-sme60" in raw_data:
        d = raw_data["bm-sme60"]
        new_bm = {
            **SME60_NEW_BENCHMARK,
            "monthlyReturns": d["monthly_filtered"],
            "stdDev": d["stddev"],
        }
        if d["ytd2026"] is not None:
            new_bm["returns"]["ytd2026"] = d["ytd2026"]
        # חשב שנתיות מחודשים זמינים
        for year in range(2020, 2026):
            annual = compute_annual(d["monthly_filtered"], year)
            if annual is not None:
                new_bm["returns"][f"y{year}"] = annual
        updated.append(new_bm)
        print(f"\n  ➕ SME60 ייווצר כמדד חדש")

    # ── שלב 4: Preview ──────────────────────────────────────────
    print("\n" + "=" * 60)
    print("📋 PREVIEW — מה עומד להיכתב ל-KV")
    print("=" * 60)

    for bm in updated:
        bm_id = bm["id"]
        mr = bm.get("monthlyReturns") or {}
        keys = sorted(mr.keys())
        old_bm = next((b for b in existing if b["id"] == bm_id), None)
        old_mr_count = len(old_bm.get("monthlyReturns") or {}) if old_bm else 0
        is_new = old_bm is None
        changed = bm_id in raw_data

        status = "🆕 חדש" if is_new else ("✏️  שונה" if changed else "   ללא שינוי")
        print(f"\n  {status} | {bm['name']} ({bm_id})")
        print(f"    currency: {bm.get('currency')} | active: {bm.get('active')}")
        print(f"    stdDev: {bm.get('stdDev')}")
        print(f"    monthlyReturns: {old_mr_count} → {len(keys)} רשומות", end="")
        if keys:
            print(f" | טווח: {keys[0]} – {keys[-1]}")
        else:
            print()
        print(f"    annual returns: {json.dumps(bm.get('returns', {}), ensure_ascii=False)}")

        if changed and bm_id in raw_data:
            # הצג דגימה של 3 ערכים ראשונים ואחרונים
            sample_keys = keys[:3] + (["..."] if len(keys) > 6 else []) + keys[-3:]
            sample = {k: round(mr[k]*100, 3) if k != "..." else "..." for k in sample_keys}
            print(f"    דוגמת ערכים (%): {sample}")

    print(f"\n  סה\"כ יועברו: {len(updated)} מדדים")

    # ── שלב 5: אישור ────────────────────────────────────────────
    print("\n" + "=" * 60)
    confirm = input("❓ לכתוב ל-KV? (כן/לא): ").strip().lower()
    if confirm not in ("כן", "yes", "y", "כ"):
        print("🛑 בוטל.")
        return

    # ── שלב 6: כתיבה ──────────────────────────────────────────
    print("\n📤 כותב benchmarks:green ל-KV...")
    result = kv_set("benchmarks:green", updated)
    print(f"  תשובת KV: {result}")

    # ── שלב 7: אימות ─────────────────────────────────────────
    print("\n🔍 מאמת...")
    verified = kv_get("benchmarks:green")
    if verified is None:
        print("  ❌ שגיאה — לא נמצא ב-KV אחרי כתיבה!")
        return

    print(f"  ✓ נמצא ב-KV | {len(verified)} מדדים")
    for bm in verified:
        mr = bm.get("monthlyReturns") or {}
        keys = sorted(mr.keys())
        print(f"    {bm['name']:25s} | {len(keys):3d} חודשים | stdDev={bm.get('stdDev')} | ytd2026={bm.get('returns',{}).get('ytd2026')}")

    print("\n✅ סיים בהצלחה!")

if __name__ == "__main__":
    main()
